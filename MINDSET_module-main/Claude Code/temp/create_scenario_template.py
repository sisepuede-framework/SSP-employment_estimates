"""
Create MINDSET Scenario Template Files

This script creates properly formatted scenario Excel files for MINDSET employment analysis.
For R users: Think of this as generating your demand shock data.frame in the right format.

Usage:
    python create_scenario_template.py
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys

# Setup paths (same as employment script)
script_dir = os.path.dirname(os.path.abspath(__file__))
mindset_root = os.path.abspath(os.path.join(script_dir, "..", ".."))
if mindset_root not in sys.path:
    sys.path.insert(0, mindset_root)
os.chdir(mindset_root)

print("="*70)
print("MINDSET SCENARIO TEMPLATE GENERATOR")
print("="*70)

# Load GLORIA product and region lists
print("\n1. Loading GLORIA product and region lists...")
try:
    products = pd.read_excel('GLORIA_template/Variables/Variable_list_MINDSET.xlsx', 'P')
    regions = pd.read_excel('GLORIA_template/Variables/Variable_list_MINDSET.xlsx', 'R')
    print(f"   ✓ Loaded {len(products)} products")
    print(f"   ✓ Loaded {len(regions)} regions")
except Exception as e:
    print(f"   ✗ Error loading GLORIA data: {e}")
    sys.exit(1)

# Save product and region lists for reference
print("\n2. Saving reference lists...")
output_dir = "Claude Code/temp"

products[['Lfd_Nr', 'Product_name', 'Product_code']].to_excel(
    f"{output_dir}/GLORIA_Products_List.xlsx", index=False
)
print(f"   ✓ Saved: {output_dir}/GLORIA_Products_List.xlsx")

regions[['Region_acronyms', 'Region_names']].to_excel(
    f"{output_dir}/GLORIA_Regions_List.xlsx", index=False
)
print(f"   ✓ Saved: {output_dir}/GLORIA_Regions_List.xlsx")

# Create scenario template
print("\n3. Creating scenario template...")

# Template header (rows 1-14)
header_data = pd.DataFrame({
    'Info': [
        'MINDSET Scenario File',
        'For employment impact analysis',
        '',
        'Instructions:',
        '1. Fill in rows 16+ with your demand shocks',
        '2. Required columns (must be named exactly):',
        '   - Producing country ISO*',
        '   - Consuming country ISO*',
        '   - Product code*',
        '   - FD code*',
        '   - Value*',
        '   - Type*',
        '',
        'Column headers below (row 15):'
    ]
})

# Column headers (row 15)
columns = [
    'Producing country ISO*',
    'Consuming country ISO*',
    'Product code*',
    'FD code*',
    'Value*',
    'Type*',
    'Notes (optional)'
]

# Example data (rows 16+)
examples = pd.DataFrame({
    'Producing country ISO*': ['USA', 'CHN', 'USA', 'USA'],
    'Consuming country ISO*': ['USA', 'CHN', 'USA', 'USA'],
    'Product code*': ['56', '1-10', '1,5,10', 'ALL'],
    'FD code*': ['FD_4', 'FD_4', 'FD_4', 'FD_4'],
    'Value*': [100000000, 10000000, 3000000, 50000000],
    'Type*': ['abs-b', 'abs-b', 'abs-b', 'abs-b'],
    'Notes (optional)': [
        '$100M to Construction in USA',
        '$10M spread across products 1-10 in China',
        '$3M to products 1,5,10 in USA',
        '$50M spread across all products in USA'
    ]
})

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Final demand"

# Write header info (rows 1-14)
for i, text in enumerate(header_data['Info'], 1):
    ws[f'A{i}'] = text
    ws[f'A{i}'].font = Font(bold=True if i in [1,4] else False)

# Write column headers (row 15)
for col_idx, col_name in enumerate(columns, 1):
    cell = ws.cell(row=15, column=col_idx)
    cell.value = col_name
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")

# Write example data (row 16+)
for row_idx, row_data in examples.iterrows():
    for col_idx, col_name in enumerate(columns, 1):
        if col_name in examples.columns:
            ws.cell(row=16+row_idx, column=col_idx).value = row_data[col_name]

# Adjust column widths
for col_idx, col_name in enumerate(columns, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

# Add instructions sheet
ws_instructions = wb.create_sheet("Instructions")
instructions_text = [
    ["MINDSET Scenario File Instructions"],
    [""],
    ["Purpose:", "Specify demand shocks for employment impact analysis"],
    [""],
    ["Required Sheet:", "'Final demand' sheet with specific column names"],
    [""],
    ["Column Descriptions:"],
    ["Column", "Description", "Example"],
    ["Producing country ISO*", "Country producing the good (use ISO codes)", "USA, CHN, BRA, or ALL"],
    ["Consuming country ISO*", "Country consuming the good", "USA, CHN, BRA, or ALL"],
    ["Product code*", "GLORIA product number (1-120)", "56 or 1-10 or 1,5,10 or ALL"],
    ["FD code*", "Final demand category", "FD_4 (investment) or FD_1 (household)"],
    ["Value*", "Amount in USD", "100000000 (=$100M)"],
    ["Type*", "How to interpret value", "abs-b (absolute) or rel-b (relative)"],
    [""],
    ["Special Codes:"],
    ["Code", "Meaning"],
    ["ALL", "All countries or all products (spreads value proportionally)"],
    ["1-10", "Products 1 through 10 (hyphen = range)"],
    ["1,5,10", "Products 1, 5, and 10 (comma = list)"],
    [""],
    ["Type Options:"],
    ["Type", "Meaning"],
    ["abs-b", "Absolute value in dollars (recommended)"],
    ["rel-b", "Relative change as decimal (0.10 = 10% increase from baseline)"],
    [""],
    ["Tips:"],
    ["- Delete example rows (16-19) before creating your scenario"],
    ["- Multiple rows sum up automatically"],
    ["- Rows 1-14 are ignored (metadata), row 15 must be headers"],
    ["- Save file in: GLORIA_template/Scenarios/"],
    [""],
    ["Reference Files:"],
    ["- GLORIA_Products_List.xlsx: All 120 products with names"],
    ["- GLORIA_Regions_List.xlsx: All 162 regions with names"],
]

for row_idx, row_data in enumerate(instructions_text, 1):
    for col_idx, text in enumerate(row_data, 1):
        cell = ws_instructions.cell(row=row_idx, column=col_idx)
        cell.value = text
        if row_idx == 1 or (row_idx > 6 and row_data[0] in ["Column", "Code", "Type", "Tips:", "Reference Files:"]):
            cell.font = Font(bold=True)

for col_idx in range(1, 4):
    ws_instructions.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 30

# Save template
template_file = f"{output_dir}/Scenario_Template.xlsx"
wb.save(template_file)
print(f"   ✓ Created: {template_file}")

print("\n4. Creating example scenarios...")

# Example 1: Test each product with $1M
scenario_all_products = pd.DataFrame({
    'Producing country ISO*': ['USA'] * 120,
    'Consuming country ISO*': ['USA'] * 120,
    'Product code*': [str(i) for i in range(1, 121)],
    'FD code*': ['FD_4'] * 120,
    'Value*': [1000000] * 120,  # $1M per product
    'Type*': ['abs-b'] * 120,
    'Notes (optional)': [f'$1M to product {i}' for i in range(1, 121)]
})

wb2 = openpyxl.load_workbook(template_file)
ws2 = wb2["Final demand"]

# Clear example data
for row in range(16, 20):
    for col in range(1, 8):
        ws2.cell(row=row, column=col).value = None

# Write all products data
for row_idx, row_data in scenario_all_products.iterrows():
    for col_idx, col_name in enumerate(columns, 1):
        if col_name in scenario_all_products.columns:
            ws2.cell(row=16+row_idx, column=col_idx).value = row_data[col_name]

example_file_1 = f"GLORIA_template/Scenarios/Example_1M_Per_Product.xlsx"
wb2.save(example_file_1)
print(f"   ✓ Created: {example_file_1}")
print(f"      Test $1M investment in each of 120 products")

# Example 2: Construction sector shock
scenario_construction = pd.DataFrame({
    'Producing country ISO*': ['USA'],
    'Consuming country ISO*': ['USA'],
    'Product code*': ['56'],  # Construction
    'FD code*': ['FD_4'],
    'Value*': [100000000],  # $100M
    'Type*': ['abs-b'],
    'Notes (optional)': ['$100M infrastructure investment']
})

wb3 = openpyxl.load_workbook(template_file)
ws3 = wb3["Final demand"]

# Clear example data
for row in range(16, 20):
    for col in range(1, 8):
        ws3.cell(row=row, column=col).value = None

# Write construction data
for row_idx, row_data in scenario_construction.iterrows():
    for col_idx, col_name in enumerate(columns, 1):
        if col_name in scenario_construction.columns:
            ws3.cell(row=16+row_idx, column=col_idx).value = row_data[col_name]

example_file_2 = f"GLORIA_template/Scenarios/Example_Construction_100M.xlsx"
wb3.save(example_file_2)
print(f"   ✓ Created: {example_file_2}")
print(f"      $100M construction investment in USA")

# Example 3: Multiple countries comparison
scenario_multi_country = pd.DataFrame({
    'Producing country ISO*': ['USA', 'CHN', 'DEU', 'BRA', 'IND'],
    'Consuming country ISO*': ['USA', 'CHN', 'DEU', 'BRA', 'IND'],
    'Product code*': ['56', '56', '56', '56', '56'],
    'FD code*': ['FD_4'] * 5,
    'Value*': [100000000] * 5,
    'Type*': ['abs-b'] * 5,
    'Notes (optional)': [
        '$100M construction in USA',
        '$100M construction in China',
        '$100M construction in Germany',
        '$100M construction in Brazil',
        '$100M construction in India'
    ]
})

wb4 = openpyxl.load_workbook(template_file)
ws4 = wb4["Final demand"]

# Clear example data
for row in range(16, 20):
    for col in range(1, 8):
        ws4.cell(row=row, column=col).value = None

# Write multi-country data
for row_idx, row_data in scenario_multi_country.iterrows():
    for col_idx, col_name in enumerate(columns, 1):
        if col_name in scenario_multi_country.columns:
            ws4.cell(row=16+row_idx, column=col_idx).value = row_data[col_name]

example_file_3 = f"GLORIA_template/Scenarios/Example_MultiCountry_Comparison.xlsx"
wb4.save(example_file_3)
print(f"   ✓ Created: {example_file_3}")
print(f"      Compare $100M construction across 5 countries")

print("\n" + "="*70)
print("TEMPLATE GENERATION COMPLETE")
print("="*70)

print("\nFiles created:")
print(f"  1. {output_dir}/Scenario_Template.xlsx")
print(f"  2. {output_dir}/GLORIA_Products_List.xlsx (reference)")
print(f"  3. {output_dir}/GLORIA_Regions_List.xlsx (reference)")
print(f"  4. GLORIA_template/Scenarios/Example_1M_Per_Product.xlsx")
print(f"  5. GLORIA_template/Scenarios/Example_Construction_100M.xlsx")
print(f"  6. GLORIA_template/Scenarios/Example_MultiCountry_Comparison.xlsx")

print("\nNext steps:")
print("  1. Open Scenario_Template.xlsx")
print("  2. Delete example rows (16-19)")
print("  3. Add your demand shocks starting at row 16")
print("  4. Save as: GLORIA_template/Scenarios/YourScenarioName.xlsx")
print("  5. Update line 103 in RunMINDSET_EmploymentOnly.py:")
print("     scenario_name = 'YourScenarioName'")
print("  6. Run the employment script!")

print("\nOr use one of the example files to test first:")
print("  scenario_name = 'Example_Construction_100M'")
print()
