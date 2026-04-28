"""
Generate Custom MINDSET Scenarios Programmatically

For R users: Think of this as creating your demand shock data.frame in R,
then saving it to Excel format for MINDSET.

Usage Examples:
    # Generate $1M shock to each product
    python generate_custom_scenario.py --mode all_products --amount 1000000 --country USA

    # Generate shock to specific products
    python generate_custom_scenario.py --mode specific_products --products "1,5,10,56" --amount 10000000 --country USA

    # Generate shock to product range
    python generate_custom_scenario.py --mode product_range --range "1-20" --amount 5000000 --country CHN
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys
import argparse

# Setup paths
script_dir = os.path.dirname(os.path.abspath(__file__))
mindset_root = os.path.abspath(os.path.join(script_dir, "..", ".."))
if mindset_root not in sys.path:
    sys.path.insert(0, mindset_root)
os.chdir(mindset_root)

def create_scenario_workbook(scenario_data, scenario_name, notes=""):
    """
    Create a properly formatted MINDSET scenario Excel file.

    Parameters:
    -----------
    scenario_data : pd.DataFrame
        Must have columns: Producing country ISO*, Consuming country ISO*,
                          Product code*, FD code*, Value*, Type*
    scenario_name : str
        Name of the scenario (will be used as filename)
    notes : str
        Description of the scenario

    R equivalent:
    -------------
    create_scenario_workbook <- function(scenario_data, scenario_name, notes="") {
        # Create Excel workbook with proper MINDSET structure
        # Save to GLORIA_template/Scenarios/
    }
    """

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Final demand"

    # Header info (rows 1-14)
    header_text = [
        f"MINDSET Scenario: {scenario_name}",
        f"Description: {notes}",
        "",
        "This file specifies demand shocks for employment impact analysis.",
        "",
        "Structure:",
        "- Rows 1-14: Metadata (this section)",
        "- Row 15: Column headers",
        "- Rows 16+: Demand shock data",
        "",
        "DO NOT modify rows 1-14 or row 15 structure.",
        "Only edit data rows (16+).",
        "",
        "Column headers below:"
    ]

    for i, text in enumerate(header_text, 1):
        ws[f'A{i}'] = text
        if i in [1, 2]:
            ws[f'A{i}'].font = Font(bold=True, size=12)

    # Column headers (row 15)
    columns = [
        'Producing country ISO*',
        'Consuming country ISO*',
        'Product code*',
        'FD code*',
        'Value*',
        'Type*',
        'Notes (optional)'
    ]

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=15, column=col_idx)
        cell.value = col_name
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Write data (row 16+)
    for row_idx, row_data in scenario_data.iterrows():
        for col_idx, col_name in enumerate(columns, 1):
            if col_name in scenario_data.columns:
                ws.cell(row=16+row_idx, column=col_idx).value = row_data[col_name]

    # Adjust column widths
    for col_idx in range(1, len(columns)+1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 25

    # Save
    output_file = f"GLORIA_template/Scenarios/{scenario_name}.xlsx"
    wb.save(output_file)

    return output_file

def generate_all_products_scenario(amount, country):
    """
    R equivalent:
    generate_all_products_scenario <- function(amount, country) {
        products <- 1:120
        data.frame(
            country_prod = rep(country, 120),
            country_cons = rep(country, 120),
            product = products,
            fd_code = rep("FD_4", 120),
            value = rep(amount, 120),
            type = rep("abs-b", 120)
        )
    }
    """
    scenario_data = pd.DataFrame({
        'Producing country ISO*': [country] * 120,
        'Consuming country ISO*': [country] * 120,
        'Product code*': [str(i) for i in range(1, 121)],
        'FD code*': ['FD_4'] * 120,
        'Value*': [amount] * 120,
        'Type*': ['abs-b'] * 120,
        'Notes (optional)': [f'${amount:,} to product {i}' for i in range(1, 121)]
    })
    return scenario_data

def generate_specific_products_scenario(products_list, amount, country):
    """
    R equivalent:
    generate_specific_products_scenario <- function(products_list, amount, country) {
        n <- length(products_list)
        data.frame(
            country_prod = rep(country, n),
            country_cons = rep(country, n),
            product = products_list,
            fd_code = rep("FD_4", n),
            value = rep(amount, n),
            type = rep("abs-b", n)
        )
    }
    """
    scenario_data = pd.DataFrame({
        'Producing country ISO*': [country] * len(products_list),
        'Consuming country ISO*': [country] * len(products_list),
        'Product code*': [str(p) for p in products_list],
        'FD code*': ['FD_4'] * len(products_list),
        'Value*': [amount] * len(products_list),
        'Type*': ['abs-b'] * len(products_list),
        'Notes (optional)': [f'${amount:,} to product {p}' for p in products_list]
    })
    return scenario_data

def generate_product_range_scenario(start, end, amount, country):
    """Generate scenario for a range of products."""
    products_list = list(range(start, end+1))
    return generate_specific_products_scenario(products_list, amount, country)

def generate_multi_country_scenario(countries_list, product, amount):
    """
    R equivalent:
    generate_multi_country_scenario <- function(countries_list, product, amount) {
        n <- length(countries_list)
        data.frame(
            country_prod = countries_list,
            country_cons = countries_list,
            product = rep(product, n),
            fd_code = rep("FD_4", n),
            value = rep(amount, n),
            type = rep("abs-b", n)
        )
    }
    """
    scenario_data = pd.DataFrame({
        'Producing country ISO*': countries_list,
        'Consuming country ISO*': countries_list,
        'Product code*': [str(product)] * len(countries_list),
        'FD code*': ['FD_4'] * len(countries_list),
        'Value*': [amount] * len(countries_list),
        'Type*': ['abs-b'] * len(countries_list),
        'Notes (optional)': [f'${amount:,} to product {product} in {c}' for c in countries_list]
    })
    return scenario_data

def main():
    parser = argparse.ArgumentParser(description='Generate MINDSET scenario files')
    parser.add_argument('--mode', required=True,
                       choices=['all_products', 'specific_products', 'product_range', 'multi_country'],
                       help='Type of scenario to generate')
    parser.add_argument('--amount', type=float, required=True,
                       help='Investment amount in USD')
    parser.add_argument('--country', type=str, default='USA',
                       help='Country ISO code (default: USA)')
    parser.add_argument('--products', type=str,
                       help='Comma-separated product codes (e.g., "1,5,10,56")')
    parser.add_argument('--range', type=str,
                       help='Product range (e.g., "1-20")')
    parser.add_argument('--countries', type=str,
                       help='Comma-separated country codes (e.g., "USA,CHN,DEU")')
    parser.add_argument('--product', type=int,
                       help='Single product code for multi-country scenario')
    parser.add_argument('--name', type=str,
                       help='Custom scenario name (optional)')

    args = parser.parse_args()

    print("="*70)
    print("MINDSET CUSTOM SCENARIO GENERATOR")
    print("="*70)

    # Generate scenario data based on mode
    if args.mode == 'all_products':
        print(f"\nGenerating: ${args.amount:,.0f} to each of 120 products in {args.country}")
        scenario_data = generate_all_products_scenario(args.amount, args.country)
        scenario_name = args.name or f"AllProducts_{int(args.amount/1e6)}M_{args.country}"
        notes = f"${args.amount:,.0f} investment in each of 120 GLORIA products in {args.country}"

    elif args.mode == 'specific_products':
        if not args.products:
            print("Error: --products required for specific_products mode")
            sys.exit(1)
        products_list = [int(p.strip()) for p in args.products.split(',')]
        print(f"\nGenerating: ${args.amount:,.0f} to products {products_list} in {args.country}")
        scenario_data = generate_specific_products_scenario(products_list, args.amount, args.country)
        scenario_name = args.name or f"Products_{'_'.join(map(str, products_list))}_{int(args.amount/1e6)}M_{args.country}"
        notes = f"${args.amount:,.0f} investment in products {products_list} in {args.country}"

    elif args.mode == 'product_range':
        if not args.range:
            print("Error: --range required for product_range mode")
            sys.exit(1)
        start, end = map(int, args.range.split('-'))
        print(f"\nGenerating: ${args.amount:,.0f} to products {start}-{end} in {args.country}")
        scenario_data = generate_product_range_scenario(start, end, args.amount, args.country)
        scenario_name = args.name or f"Products_{start}to{end}_{int(args.amount/1e6)}M_{args.country}"
        notes = f"${args.amount:,.0f} investment in products {start}-{end} in {args.country}"

    elif args.mode == 'multi_country':
        if not args.countries or args.product is None:
            print("Error: --countries and --product required for multi_country mode")
            sys.exit(1)
        countries_list = [c.strip() for c in args.countries.split(',')]
        print(f"\nGenerating: ${args.amount:,.0f} to product {args.product} in countries {countries_list}")
        scenario_data = generate_multi_country_scenario(countries_list, args.product, args.amount)
        scenario_name = args.name or f"Product{args.product}_{'_'.join(countries_list)}_{int(args.amount/1e6)}M"
        notes = f"${args.amount:,.0f} investment in product {args.product} across {len(countries_list)} countries"

    # Create Excel file
    output_file = create_scenario_workbook(scenario_data, scenario_name, notes)

    print(f"\n✓ Created scenario file: {output_file}")
    print(f"✓ Total rows: {len(scenario_data)}")
    print(f"✓ Total investment: ${scenario_data['Value*'].sum():,.0f}")

    print("\nTo run this scenario:")
    print(f"  1. Open RunMINDSET_EmploymentOnly.py")
    print(f"  2. Set line 103 to: scenario_name = '{scenario_name}'")
    print(f"  3. Run the script!")

    print("\n" + "="*70)

if __name__ == "__main__":
    # If no arguments, show help and examples
    if len(sys.argv) == 1:
        print("="*70)
        print("MINDSET CUSTOM SCENARIO GENERATOR")
        print("="*70)
        print("\nUsage examples:\n")

        print("1. Generate $1M shock to each of 120 products in USA:")
        print("   python generate_custom_scenario.py --mode all_products --amount 1000000 --country USA\n")

        print("2. Generate $10M shock to specific products (1, 5, 10, 56) in China:")
        print("   python generate_custom_scenario.py --mode specific_products --products '1,5,10,56' --amount 10000000 --country CHN\n")

        print("3. Generate $5M shock to products 1-20 in Germany:")
        print("   python generate_custom_scenario.py --mode product_range --range '1-20' --amount 5000000 --country DEU\n")

        print("4. Generate $100M shock to construction (56) in 5 countries:")
        print("   python generate_custom_scenario.py --mode multi_country --countries 'USA,CHN,DEU,BRA,IND' --product 56 --amount 100000000\n")

        print("5. Custom name:")
        print("   python generate_custom_scenario.py --mode all_products --amount 1000000 --country USA --name 'MyCustomScenario'\n")

        print("For detailed help:")
        print("   python generate_custom_scenario.py --help\n")
    else:
        main()
