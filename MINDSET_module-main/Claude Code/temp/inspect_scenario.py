#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Inspect scenario file structure
"""

import pandas as pd
import openpyxl

scenario_file = r"C:\Users\festeves\OneDrive - RAND Corporation\Courses\Dissertation\3rd Paper\Mindset - WB\MINDSET_module-main\MINDSET_module-main\GLORIA_template\Scenarios\New template.xlsx"

print("=" * 80)
print("SCENARIO FILE STRUCTURE ANALYSIS")
print("=" * 80)

# Load workbook to see sheet names
wb = openpyxl.load_workbook(scenario_file, read_only=True, data_only=True)

print(f"\n1. SHEET NAMES:")
for i, sheet_name in enumerate(wb.sheetnames):
    print(f"   Sheet {i+1}: '{sheet_name}'")

print(f"\n2. INSPECTING EACH SHEET:")
print("-" * 80)

for sheet_name in wb.sheetnames:
    print(f"\n   SHEET: {sheet_name}")
    print(f"   " + "-" * 76)

    # Read sheet with pandas
    try:
        df = pd.read_excel(scenario_file, sheet_name=sheet_name, nrows=10)
        print(f"   Dimensions: {df.shape[0]} rows × {df.shape[1]} columns (sample)")
        print(f"   Columns: {list(df.columns[:5])}..." if len(df.columns) > 5 else f"   Columns: {list(df.columns)}")
        print(f"\n   First 3 rows:")
        print("   " + df.head(3).to_string().replace('\n', '\n   '))
    except Exception as e:
        print(f"   Error reading: {e}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
