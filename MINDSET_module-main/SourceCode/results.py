# -*- coding: utf-8 -*-
"""
Created on Wed Aug  2 10:57:33 2023

@author: wb582890
"""

import os
import pandas as pd
import numpy as np
import pyarrow as pa
import pyarrow.parquet as pq
from scipy import sparse
import pickle
import openpyxl

class results:
    
    def __init__(self, IO_PATH = os.getcwd() + "\\", scenario="BA", country_results=0, iter_results=0, regions=None):
        self.IO_PATH = IO_PATH
        self.country_results = country_results
        self.iter_results = iter_results
        self.scenario = scenario
        self.REGIONS = regions

        self.BASELINE_RESIDUALS_PATH = "GLORIA_results\\baseline_residuals.xlsx"
    
    def save_change(self, tax_cou, results_list, iter_list=None, year=None):
        [output_change, empl_change, gdp_change, hh_change, dy_df, price_change, tax_revenue, emissions] = results_list
        if year:
            header_ = (year == 2019)

            with (pd.ExcelWriter(self.IO_PATH + 'GLORIA_results\\FullResults_' + self.scenario + '.xlsx',engine="openpyxl", mode='w') if year == 2019 else
                  pd.ExcelWriter(self.IO_PATH + 'GLORIA_results\\FullResults_' + self.scenario + '.xlsx',engine="openpyxl", mode='a', if_sheet_exists="overlay")) as ResultsWriter:
                book = ResultsWriter.sheets
                output_change.assign(year=year).to_excel(ResultsWriter, sheet_name='output', startrow=(0 if year == 2019 else book['output'].max_row), header=header_)
                empl_change.assign(year=year).to_excel(ResultsWriter, sheet_name='employment', startrow=(0 if year == 2019 else book['employment'].max_row), header=header_)
                gdp_change.assign(year=year).to_excel(ResultsWriter, sheet_name='gdp', startrow=(0 if year == 2019 else book['gdp'].max_row), header=header_)
                dy_df.assign(year=year).to_excel(ResultsWriter, sheet_name='demand', startrow=(0 if year == 2019 else book['demand'].max_row), header=header_)
                hh_change.assign(year=year).to_excel(ResultsWriter, sheet_name='household', startrow=(0 if year == 2019 else book['household'].max_row), header=header_)
                price_change.assign(year=year).to_excel(ResultsWriter, sheet_name='price', startrow=(0 if year == 2019 else book['price'].max_row), header=header_)
                tax_revenue.assign(year=year).to_excel(ResultsWriter, sheet_name='revenue', startrow=(0 if year == 2019 else book['revenue'].max_row), header=header_)
                emissions.assign(year=year).to_excel(ResultsWriter, sheet_name='emissions', startrow=(0 if year == 2019 else book['emissions'].max_row), header=header_)
        else:
            with pd.ExcelWriter(
                    self.IO_PATH + 'GLORIA_results\\FullResults_' + self.scenario + '.xlsx',
                    engine="xlsxwriter") as ResultsWriter:
                output_change.to_excel(ResultsWriter, sheet_name='output')
                empl_change.to_excel(ResultsWriter, sheet_name='employment')
                gdp_change.to_excel(ResultsWriter, sheet_name='gdp')
                dy_df.to_excel(ResultsWriter, sheet_name='demand')
                hh_change.to_excel(ResultsWriter, sheet_name='household')
                price_change.to_excel(ResultsWriter, sheet_name='price')
                tax_revenue.to_excel(ResultsWriter, sheet_name='revenue')
                emissions.to_excel(ResultsWriter, sheet_name='emissions')
            
    def adj_residuals(self, results_list, residuals_file, year=None):
        if residuals_file:
        # TODO: does not currently work for emission and tax revenues!!
        # TODO: because for those we don't print full matrix (i.e. 19680 items)
            [output_change, empl_change, gdp_change, hh_change, dy_df, price_change, tax_revenue, emissions] = results_list
            output_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='output').query("year == @year").drop(columns=['year'])
            empl_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='employment').query("year == @year").drop(columns=['year'])
            gdp_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='gdp').query("year == @year").drop(columns=['year'])
            hh_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='household').query("year == @year").drop(columns=['year'])
            price_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='price').query("year == @year").drop(columns=['year'])
            tax_revenue_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='revenue').query("year == @year").drop(columns=['year'])
            # emissions_residual = pd.read_excel("{}{}".format(self.IO_PATH, residuals_file), sheet_name='emissions')
            
            # TODO: remove later!! NPISH
            output_residual = output_residual.pipe(lambda d: d.loc[:, ~d.columns.str.contains("npish")])
            empl_residual = empl_residual.pipe(lambda d: d.loc[:, ~d.columns.str.contains("npish")])
            gdp_residual = gdp_residual.pipe(lambda d: d.loc[:, ~d.columns.str.contains("npish")])
            # import pdb; pdb.set_trace()

            output_adj = output_change.copy()
            output_adj.iloc[:,5:] = output_change.iloc[:,5:].values - output_residual.iloc[:,6:].values

            empl_adj = empl_change.copy()
            empl_adj.iloc[:,5:] = empl_change.iloc[:,5:].values - empl_residual.iloc[:,6:].values

            gdp_adj = gdp_change.copy()
            gdp_adj.iloc[:,5:] = gdp_change.iloc[:,5:].values - gdp_residual.iloc[:,6:].values

            hh_adj = hh_change.copy()
            hh_adj.iloc[:,5:] = hh_change.iloc[:,5:].values - hh_residual.iloc[:,6:].values

            price_adj = price_change.copy()
            price_adj.iloc[:,4:] = price_change.iloc[:,4:].values - price_residual.iloc[:,5:].values

            # tax_revenue_adj = tax_revenue.copy()
            # tax_revenue_adj.iloc[:,2:] = tax_revenue.iloc[:,2:].values - tax_revenue_residual.iloc[:,3:].values

            # emissions_adj = emissions.copy()
            # emissions_adj.iloc[:,[2,4,5]] = emissions.iloc[:,[2,4,5]].values - emissions_residual.iloc[:,[3,5,6]].values

            return [output_adj, empl_adj, gdp_adj, hh_adj, dy_df, price_adj, tax_revenue, emissions]
        else:
            return results_list

    def save_INDBASE(self, io, total_output, file):
        io_df = self.MRIO_mat_to_df(io, iso3=True)
        total_output = self.MRIO_vec_to_df(total_output, "output")
        io_df = io_df.merge(total_output, how='left', on=['target-country','target-sector'])

        io_df = io_df.rename(columns={'value':'z_bp','target-country-iso3':'REG_imp','origin-country-iso3':'REG_exp',
                             'target-sector':'PROD_COMM','origin-sector':'TRAD_COMM'})

        io_df = io_df[['REG_imp','PROD_COMM','REG_exp','TRAD_COMM','z_bp','output']].copy()
        io_df['a_bp'] = io_df['z_bp'] / io_df['output']
        io_df['a_tech'] = io_df.groupby(['REG_imp','PROD_COMM','TRAD_COMM'])['z_bp'].transform('sum') / io_df.groupby(['REG_imp','PROD_COMM'])['z_bp'].transform('sum')
        
        with open(file, 'wb') as file:
            pickle.dump(io_df, file)

    def save_io(self, io, name, append_to_file=None):
        io_df = self.MRIO_mat_to_df(io).rename(columns={'value':'value_{}'.format(name)})
        if append_to_file:
            df = pq.read_table(append_to_file).to_pandas()
            df = df.merge(io_df, how='left', on=['target-country','origin-country','target-sector','origin-sector'])
            io_df = df.copy()
        table = pa.Table.from_pandas(io_df)
        pq.write_table(table, self.IO_PATH + 'GLORIA_results\\io_' + self.scenario + '_' + name + '.parquet')
    
    def save_trajectory(self, dynamic_list):
        [dynamic_output, dynamic_empl, dynamic_gdp, dynamic_price, dynamic_revenue] = dynamic_list
        with pd.ExcelWriter(
                self.IO_PATH + 'GLORIA_results\\Trajectory_macro_' + self.scenario + '.xlsx',
                engine="xlsxwriter") as ResultsWriter:
            dynamic_output.to_excel(ResultsWriter, sheet_name='output')
            dynamic_empl.to_excel(ResultsWriter, sheet_name='employment')
            dynamic_gdp.to_excel(ResultsWriter, sheet_name='gdp')
            dynamic_price.to_excel(ResultsWriter, sheet_name='price')
            dynamic_revenue.to_excel(ResultsWriter, sheet_name='revenue')

    def MRIO_vec_to_df(self, vector, name, iso3=False):
        # vector is 19680, 120x164
        vec = pd.DataFrame(vector).reset_index()
        vec.columns = ['index', name]
        vec['target-country'] = (np.floor(vec['index'] / 120)).astype('int16') + 1
        vec['target-sector'] =  (vec['index'] % 120).astype('int16') + 1
        vec[name] = vec[name].astype('float32')
        vec = vec.drop(columns=['index'])
        return vec

    def MRIO_mat_to_df(self, mrio, iso3=False):
        # MRIO is 19680 x 19680 matrix, this is 120x164, 120x164   
        print(1)
        # 1. put into dataframe

        # convert to sparse matrix
        mrio = sparse.csc_matrix(mrio)
        print(2)
        mrio = pd.DataFrame.sparse.from_spmatrix(mrio).reset_index()

        # now we have 
        # columns: index [sec1-reg1 = 0] [sec2-reg1 = 1] ... [sec120 - reg164 = 19680]
        # rows: 1 2 3 ... 19680
        print(3)
        # 2. melt
        mrio = pd.melt(mrio, id_vars=['index'])
        print(4)
        # now it's
        # columns: index[==row-index] variable[==column-index] value
        # rows: 1 2 3 ... 19680
        mrio = mrio[mrio['value']!=0].copy()
        print(5)
        # 3. reindex
        mrio['origin-country'] = (np.floor(mrio['index'] / 120)).astype('int16') + 1
        mrio['origin-sector'] = (mrio['index'] % 120).astype('int16') + 1
        mrio['target-country'] = (np.floor(mrio['variable'] / 120)).astype('int16') + 1
        mrio['target-sector'] = (mrio['variable'] % 120).astype('int16') + 1

        if iso3:
            # 4. merge country-iso
            mrio = self.attach_iso3(mrio, self.REGIONS)
        mrio['value'] = mrio['value'].astype('float32')
        mrio = mrio.drop(columns=['index','variable'])
        
        return mrio
    
    def attach_iso3(self, df, regions):
        """ 
        Attaches ISO3 codes from the `regions` input to the dataframe
        `origin-country` and `target-country` have to be present in the dataframe
        """
        regions = regions.drop(columns=['Region_names'])
        regions['Region_acronyms'] = regions['Region_acronyms'].astype('category')
        df = df.merge(regions, how='left', left_on=['target-country'], right_on=['Lfd_Nr']).rename(columns={'Region_acronyms':'target-country-iso3'})\
            .drop(columns=['Lfd_Nr'])
        df = df.merge(regions, how='left', left_on=['origin-country'], right_on=['Lfd_Nr']).rename(columns={'Region_acronyms':'origin-country-iso3'})\
            .drop(columns=['Lfd_Nr'])
        return df
    
    